from pathlib import Path

from openpyxl import Workbook

from app.ingest.excel_parser import parse_excel_grade23
from app.ingest.excel_parser import parse_excel_grade1_student_scores


def test_parse_grade23_sample_totals():
    sample = (
        Path(__file__).resolve().parents[3]
        / "高二成绩"
        / "高二2025学年第二学期期中考试学生成绩明细表.xlsx"
    )
    result = parse_excel_grade23(str(sample), grade=2)

    assert result["kind"] == "student_scores"
    assert len(result["students"]) > 700

    totals = {
        row["total_type"]: row
        for row in result["total_scores"]
        if row["student_id"] == "7240101"
    }
    assert totals["主三门"]["total_score"] == 325.5
    assert totals["+3"]["total_score"] == 174
    assert totals["3+3"]["total_score"] == 499.5
    assert totals["3+3"]["xueji_rank"] == 291


def test_parse_grade23_sample_subjects():
    sample = (
        Path(__file__).resolve().parents[3]
        / "高二成绩"
        / "高二2025学年第二学期期中考试学生成绩明细表.xlsx"
    )
    result = parse_excel_grade23(str(sample), grade=2)

    subjects = {
        row["subject"]: row
        for row in result["subject_scores"]
        if row["student_id"] == "7240101"
    }
    assert subjects["语文"]["raw_score"] == 97
    assert subjects["语文"]["grade_percentile"] == 0.7508
    assert subjects["物理"]["raw_score"] == 48
    assert subjects["物理"]["grade_score"] == 52
    assert "生物" not in subjects


def test_parse_grade1_percentiles(tmp_path):
    workbook = Workbook()
    ws = workbook.active
    ws.title = "学生成绩明细"
    ws.append(["学生成绩（在籍）"])
    ws.append(["学号", "班级", "学籍", "姓名", "语文", None, "数学", None, "主三门", None, None, None])
    ws.append([None, None, None, None, "分数", "年级百分位", "分数", "年级百分位", "总分", "年级百分位", "学籍排名", "年级排名"])
    ws.merge_cells("A2:A3")
    ws.merge_cells("B2:B3")
    ws.merge_cells("C2:C3")
    ws.merge_cells("D2:D3")
    ws.merge_cells("E2:F2")
    ws.merge_cells("G2:H2")
    ws.merge_cells("I2:L2")
    ws.append(["7240101", "01", "1", "卞幻", 92, "45.01%", 106, "66.84%", 306, "47.88%", 283, 283])
    path = tmp_path / "2024级2024学年第二学期期中考试.xlsx"
    workbook.save(path)

    result = parse_excel_grade1_student_scores(path)

    subjects = {
        row["subject"]: row
        for row in result["subject_scores"]
        if row["student_id"] == "7240101"
    }
    totals = {
        row["total_type"]: row
        for row in result["total_scores"]
        if row["student_id"] == "7240101"
    }
    assert subjects["语文"]["grade_percentile"] == 0.4501
    assert subjects["数学"]["grade_percentile"] == 0.6684
    assert totals["主三门"]["grade_percentile"] == 0.4788
    assert totals["主三门"]["xueji_rank"] == 283
    assert totals["主三门"]["grade_rank"] == 283


def test_parse_grade23_class_average_workbook(tmp_path):
    workbook = Workbook()
    ws = workbook.active
    ws.title = "班级均分"
    ws.append(
        [
            "班型",
            "班级",
            "班主任",
            "语文",
            "数学",
            "英语",
            "物理",
            None,
            "化学",
            None,
            "加3同均分",
            "主三门",
            "3+3总分",
        ]
    )
    ws.append([None, None, None, None, None, None, "原始", "等级", "原始", "等级", None, None, None])
    ws.append(["平行班", "01", "张老师", 101.2, 108.3, 110.4, 55.1, 61.2, 58.3, 62.4, 180.5, 319.9, 500.4])
    ws.append([None, "02", "李老师", 99.1, 103.2, 109.3, 53.1, 59.2, 57.3, 60.4, 176.5, 311.6, 488.1])
    path = tmp_path / "高二2025学年第二学期期中考试班级均分表.xlsx"
    workbook.save(path)

    result = parse_excel_grade23(str(path), grade=2)

    assert result["kind"] == "class_averages"
    assert len(result["class_averages"]) == 2
    first = result["class_averages"][0]
    assert first["class_type"] == "平行班"
    assert first["class_num"] == 1
    assert first["teacher_name"] == "张老师"
    assert first["subject_averages"]["语文"] == 101.2
    assert first["subject_averages"]["物理_等级"] == 61.2
    assert first["total_averages"]["+3"] == 180.5
    assert first["total_averages"]["主三门"] == 319.9
    assert first["total_averages"]["3+3"] == 500.4
